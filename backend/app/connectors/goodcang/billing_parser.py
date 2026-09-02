"""谷仓账单 xlsx 费用明细解析器。

把 billing_export 返回的 base64 xlsx 解析为费用行明细，并归类到五大类：
storage(仓储) / inbound(入库) / outbound(出库操作) / transport(运输) / other(其他)

谷仓账单 xlsx 格式 **高度多变**（按账单/主体/时期不同而异），本解析器做了容错适配：

1. sheet 名不固定：可能是中文（入库费用明细/仓租费用明细…），也可能是英文
   单数（Inbound Fee/Storage Fee/Order Fee…）或复数（Inbound Fees/Order Fees/…），
   还可能是长名（Warehouse Storage Service Fees）。因此按**关键字模糊匹配**：
   - 含 "Inbound" / "入库" → 入库费 sheet
   - 含 "Storage" / "仓租" / "仓储" → 仓储费 sheet
   - 含 "Order" / "订单" → 订单费 sheet
   - 含 "Return" / "退货" → 退货费 sheet
   - 含 "VAS" / "增值" → 增值服务 sheet

2. 列名不固定：中文（费用类型/费用明细/本期费用小计/币种）或英文
   （Fee Type / Fee Details|Fee Detail / Total Charges / Currency）。

3. 金额列：英文格式用 **Total Charges**（已扣减 Recharge/Refund），中文用 本期费用小计。

归类规则：把「费用类型 + 费用明细」文本拼接后按关键字匹配（先命中先生效）：
- storage：仓租/仓储/超库龄/Storage
- inbound：卸货/入库/Inbound Handling/Inbound
- outbound：操作/打包/退件处理/质检/理货/Handling/Pick/Pack
- transport：运费/道路通行/燃油附加/运输/Freight/Shipping/Postage/Delivery/
             Road Toll/Fuel Surcharge/Peak Season/Address correction
- 其余 → other
"""
from __future__ import annotations

import base64
import io
from decimal import Decimal

import openpyxl

# sheet 名关键字 → 归类兜底（用于无法从明细文本判断时；实际归类主要看明细文本）
SHEET_KEYWORDS = (
    ("inbound", ("Inbound", "入库")),
    ("storage", ("Storage", "仓租", "仓储")),
    ("order", ("Order", "订单")),
    ("return", ("Return", "退货")),
    ("vas", ("VAS", "增值")),
)

# 列名候选（中英文），按顺序匹配
FEE_TYPE_COLS = ("费用类型", "Fee Type")
FEE_NAME_COLS = ("费用明细", "Fee Details", "Fee Detail")
# 金额列：优先不含税（Tax Excl.），与账单主表总额对账一致
AMOUNT_COLS = ("本期费用小计", "Total Charges(Tax Excl.)", "Total Charges")
CURRENCY_COLS = ("币种", "Currency")

# 归类关键字 → 五大类（按顺序匹配，先命中先生效）
CATEGORY_RULES = [
    (("仓租", "仓储", "超库龄", "Storage", "storage"), "storage"),
    (("卸货", "入库操作", "入库处理", "入库费", "Inbound Handling", "Inbound"), "inbound"),
    (("出库操作", "打包费", "退件处理", "质检", "理货", "操作费", "Handling", "Pick", "Pack"), "outbound"),
    (("退件运费", "道路通行", "燃油附加", "运输费", "运费",
      "Freight", "Shipping", "Postage", "Delivery", "Road Toll",
      "Fuel Surcharge", "Peak Season", "Address correction"), "transport"),
]


def classify_fee(name: str) -> str:
    """按费用明细/类型文本归类到五大类。"""
    for kws, cat in CATEGORY_RULES:
        for kw in kws:
            if kw in name:
                return cat
    return "other"


def _col(header, candidates: tuple[str, ...]) -> int:
    """在表头中查找候选列名（中英文），返回列索引。

    支持精确匹配 + 前缀匹配（例如 "Total Charges" 前缀能匹配
    "Total Charges(Tax Excl.)"）。
    """
    # 先精确匹配
    for i, h in enumerate(header):
        if h is None:
            continue
        s = str(h).strip()
        if any(s == c for c in candidates):
            return i
    # 再前缀匹配（用于 "Total Charges(Tax Excl.)" 这类变体）
    for i, h in enumerate(header):
        if h is None:
            continue
        s = str(h).strip()
        if any(s.startswith(c) for c in candidates):
            return i
    return -1


def _cell(row, i: int):
    if i < 0 or i >= len(row):
        return None
    return row[i]


def _is_fee_sheet(sheet_name: str) -> bool:
    """判断 sheet 是否是费用明细 sheet（按关键字模糊匹配）。"""
    for _, kws in SHEET_KEYWORDS:
        for kw in kws:
            if kw.lower() in sheet_name.lower():
                return True
    # 中文关键字
    for kw in ("入库", "仓租", "仓储", "订单", "退货", "增值", "费用明细"):
        if kw in sheet_name:
            return True
    return False


def parse_billing_xlsx(b64: str) -> list[dict]:
    """解析 base64 xlsx，返回费用明细行列表。

    每行：{"fee_name": 费用明细, "fee_type": 费用类型, "amount": Decimal,
           "currency": 币种, "sheet": 来源页签}
    """
    raw = base64.b64decode(b64)
    # 注意：不能用 read_only=True，否则谷仓 xlsx 的表头行会被截断（列维度不完整）
    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    rows: list[dict] = []

    for sheet_name in wb.sheetnames:
        if not _is_fee_sheet(sheet_name):
            continue
        ws = wb[sheet_name]
        header = None
        idx = {}
        for r in ws.iter_rows(values_only=True):
            if header is None:
                header = r  # 第一行是表头
                idx = {
                    "fee_type": _col(header, FEE_TYPE_COLS),
                    "fee_name": _col(header, FEE_NAME_COLS),
                    "amount": _col(header, AMOUNT_COLS),
                    "currency": _col(header, CURRENCY_COLS),
                }
                # 如果找不到费用明细列或金额列，跳过这个 sheet
                if idx["fee_name"] < 0 or idx["amount"] < 0:
                    break
                continue
            if r is None or all(v is None for v in r):
                continue
            name = _cell(r, idx["fee_name"])
            if not name:
                continue
            amt = _cell(r, idx["amount"])
            if amt is None:
                continue
            try:
                amount = Decimal(str(amt))
            except Exception:  # noqa: BLE001
                continue
            if amount == 0:
                continue  # 跳过零金额行
            fee_type = _cell(r, idx["fee_type"]) or ""
            rows.append({
                "fee_name": str(name),
                "fee_type": str(fee_type),
                "amount": amount,
                "currency": _cell(r, idx["currency"]) or "EUR",
                "sheet": sheet_name,
            })
    wb.close()
    return rows
